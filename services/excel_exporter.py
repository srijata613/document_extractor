from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from models.document import (
    Document,
    MappedField,
)
from utils.logger import app_logger


class ExcelExportError(Exception):
    """Raised when Excel export fails."""


class ExcelExporter:

    """
    Production Excel exporter.
    """

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
    )

    THIN_BORDER = Border(

        left=Side(style="thin"),

        right=Side(style="thin"),

        top=Side(style="thin"),

        bottom=Side(style="thin"),
    )

    HEADER_ALIGNMENT = Alignment(

        horizontal="center",

        vertical="center",
    )

    CELL_ALIGNMENT = Alignment(

        vertical="top",

        wrap_text=True,
    )

    def export_document(
        self,
        document: Document,
        mapped_fields: dict[str, MappedField],
        output_path: Path,
    ) -> Path:

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Extracted Data"

        self._write_header(
            sheet,
            mapped_fields,
        )

        self._write_document(
            sheet,
            document,
            mapped_fields,
        )

        self._format_sheet(
            sheet,
        )

        self._autosize_columns(
            sheet,
        )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        try:
            workbook.save(
                output_path,
            )
        except PermissionError as exc:
            raise ExcelExportError(
                "Please close the Excel file before exporting again."
            ) from exc
        except Exception as exc:
            raise ExcelExportError(
                f"Failed to save Excel: {exc}"
            ) from exc

        app_logger.success(
            f"Excel exported to {output_path}"
        )

        return output_path

    def export_batch(
        self,
        documents: Iterable[
            tuple[
                Document,
                dict[str, MappedField],
            ]
        ],
        output_path: Path,
    ) -> Path:

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Extracted Data"

        first = True

        row = 2

        for document, mapped in documents:

            if first:

                self._write_header(
                    sheet,
                    mapped,
                )

                first = False

            self._write_document(
                sheet,
                document,
                mapped,
                row=row,
            )

            row += 1

        self._format_sheet(
            sheet,
        )

        self._autosize_columns(
            sheet,
        )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        try:
            workbook.save(
                output_path
            )

        except PermissionError as exc:
            raise ExcelExportError(
                "Please close the Excel file before exporting again."
            ) from exc

        except Exception as exc:
            raise ExcelExportError(
                f"Failed to save Excel: {exc}"
            ) from exc

        return output_path

    @staticmethod
    def _write_header(
        sheet,
        mapped_fields: dict[str, MappedField],
    ) -> None:

        column = 1

        for field_name in mapped_fields:

            cell = sheet.cell(
                row=1,
                column=column,
            )

            cell.value = field_name

            column += 1

    @staticmethod
    def _write_document(
        sheet,
        document: Document,
        mapped_fields: dict[str, MappedField],
        row: int = 2,
    ) -> None:
        """
        Write one document as one Excel row.
        """

        column = 1

        for field in mapped_fields.values():

            value = field.value

            if value is None:
                value = ""

            sheet.cell(
                row=row,
                column=column,
                value=str(value),
            )

            column += 1
            
    @staticmethod
    def sanitize_filename(
        filename: str,
    ) -> str:

        invalid = '<>:"/\\|?*'

        for ch in invalid:
            filename = filename.replace(
                ch,
                "_",
            )

        return filename.strip()
    
    @staticmethod
    def create_output_path(
        output_dir: Path,
        filename: str,
    ) -> Path:

        filename = ExcelExporter.sanitize_filename(
            filename
        )

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        output_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

        return output_dir / filename
    
    def export(
        self,
        document: Document,
        mapped_fields: dict[str, MappedField],
        output_directory: Path,
        filename: str,
    ) -> Path:

        output = self.create_output_path(
            output_directory,
            filename,
        )

        return self.export_document(
            document,
            mapped_fields,
            output,
        )
        
    def _format_sheet(
        self,
        sheet,
    ) -> None:

        for cell in sheet[1]:

            cell.fill = self.HEADER_FILL

            cell.font = self.HEADER_FONT

            cell.border = self.THIN_BORDER

            cell.alignment = self.HEADER_ALIGNMENT

        for row in sheet.iter_rows(
            min_row=2,
        ):

            for cell in row:

                cell.border = self.THIN_BORDER

                cell.alignment = self.CELL_ALIGNMENT

        sheet.freeze_panes = "A2"

        sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _autosize_columns(
        sheet,
    ) -> None:
        for column_cells in sheet.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0

                for cell in column_cells
            )

            letter = get_column_letter(
                column_cells[0].column
            )
            sheet.column_dimensions[
                letter
            ].width = min(
                max(length + 3, 15),
                60,
            )
            
excel_exporter = ExcelExporter()